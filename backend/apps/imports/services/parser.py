"""Tolerant parser: xlsx/csv -> list of raw row dicts. Never validates, never crashes.

Values are preserved verbatim (including trailing spaces like 'rohan ') so the
detectors can see the mess. Dates from xlsx cells are emitted as ISO strings.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook

COLUMNS = ["date", "description", "paid_by", "amount", "currency",
           "split_type", "split_with", "split_details", "notes"]

MAX_ROWS = 10_000  # cap staged rows so a huge sheet can't exhaust memory / CPU (O(n^2) dedup)


class ParseError(ValueError):
    """Raised when a file is too large/long to import safely."""


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def parse(file_bytes: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    return _parse_csv(file_bytes)


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    # read_only streams rows instead of materialising the whole sheet in memory
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        it = wb.worksheets[0].iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return []
        header = [(_cell_to_str(c)).strip().lower() for c in header_row]
        out = []
        for i, raw in enumerate(it, start=1):
            if i > MAX_ROWS:
                raise ParseError(f"spreadsheet has too many rows (max {MAX_ROWS:,})")
            record = {header[j] if j < len(header) else f"col{j}": _cell_to_str(v)
                      for j, v in enumerate(raw)}
            # keep only the known columns (in order), defaulting missing to ""
            record = {col: record.get(col, "") for col in COLUMNS}
            out.append({"row_number": i, "raw": record})
        return out
    finally:
        wb.close()


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for i, row in enumerate(reader, start=1):
        if i > MAX_ROWS:
            raise ParseError(f"file has too many rows (max {MAX_ROWS:,})")
        record = {col: (row.get(col) or "") for col in COLUMNS}
        out.append({"row_number": i, "raw": record})
    return out
